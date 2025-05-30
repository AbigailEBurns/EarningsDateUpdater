#import necessary stuff
import logging
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import re
from datetime import datetime, timedelta

logging.basicConfig(level=logging.DEBUG)

#global variables
url = 'https://www.zacks.com/stock/research/{}/earnings-calendar?icid=quote-stock_overview-quote_nav_tracking-zcom-left_subnav_quote_navbar-earnings_dates_announcements'
top_date_path = '//*[@id="right_content"]/section[2]/div[1]/table/tbody/tr/th'
bottom_date_path = '//*[@id="earnings_announcements_earnings_table"]/tbody/tr[1]/th'

def main():
    
    #load excel sheet
    try:
        workbook = load_workbook(filename = 'test_stocks.xlsx')
    except Exception as e:
        print(f'Unable to open: {e}')
        exit()
    
    #defines sheet as the open excel file
    sheet = workbook.active

    #iterate row by row through the sheet starting at the third row
    for row in sheet.iter_rows(min_row = 3):
        
        #get the row number from the first cell
        row_num = row[0].row
        
        #alternate between a and b columns and save the stock symbol as a variable for each:
        stock_cellA = 'A'
        date_cellA = 'C' 

        stock_cellB = 'B'
        date_cellB = 'D'

        search_stock(sheet, stock_cellA, date_cellA, row_num)

        search_stock(sheet, stock_cellB, date_cellB, row_num)
    try:
        workbook.save('test_stocksa.xlsx')
    except Exception as e:
        print(f'Failed to save workbook: {e}')

def search_stock(sheet, stock_cell, date_cell, row):

    #get the name of the stock
    stock = sheet[f'{stock_cell}{row}'].value
    
    #check if the date cell is empty
    if sheet[f'{date_cell}{row}'].value is None:
        #proceed with the process
        stock_date = stock_process(stock)

        #insert selected date into excel
        sheet[f'{date_cell}{row}'] = stock_date

        #determine what style needs to be applied to the cell and apply it
        apply_style(stock_date, sheet[f'{date_cell}{row}'])

def stock_process(stock):
    
    #call scraping function, pass symbol
    date1, date2 = scrape(stock)

    #if scrape fails to load page or elements on page
    if date1 == 'LOAD ERROR' or date2 == 'LOAD ERROR':
        return 'LOAD ERROR'
    
    #call fucntion to select which date will be passes to excel, pass dates from scraping
    stock_date = select_date(date1, date2)
    
    return stock_date

#function for scraping
def scrape(stock):

    driver = set_webdriver()

    date1, date2 = None, None
    
    #acess page
    try:
        #combines the url with the extracted stock
        driver.get(url.format(stock))

        #calls get_date1 func to get the top date
        date1 = get_date(driver, top_date_path)

        #calls get_date2 fun to get the bottom date
        date2 = get_date(driver, bottom_date_path)
        
    #excpet if it doesnt work display an error
    except Exception as e:
        print(f'ERROR scrape {stock}: {e}')
        return 'LOAD ERROR'
    
    #quit the driver when done
    finally:
        driver.quit()

    #these should be returned as a tuple containing 2 strings
    return date1, date2

#selenium webdriver
def set_webdriver():
    
    options = webdriver.FirefoxOptions()
    options.binary_location = r'C:\Program Files\Mozilla Firefox\firefox.exe'
    options.add_argument('-headless') #headless mode

    #user agent to mimic regular browser
    options.set_preference("general.useragent.override", 
                           "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:91.0) Gecko/20100101 Firefox/91.0")

    #disable webdriver flag to make the automation less detectable
    options.set_preference('dom.webdriver.enabled', False)
    options.set_preference('useAutomationExtension', False)

    #set path to webdriver service
    service = webdriver.FirefoxService(r'C:\Program Files (x84)\WebDrivers\geckodriver.exe')

    #initialize driver
    driver = webdriver.Firefox(service=service, options=options)

    #remove navigator.webdriver property to reduce visability 
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

    #wait for elements on the page to load
    driver.implicitly_wait(10)
    
    return driver

#get the date from the webpage
def get_date(driver, date_path):

    #wait for date element to load
    try:
        text = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, date_path))
            )
    except Exception as e:
            print(f'ERROR get_date: {e}')
            return None

    #call the extract date function to seperate the date from the rest of the text
    date = extract_date(driver, text)
    
    return date

#seperate a date from other text
def extract_date(driver, text):

    #the date pattern to look for as a regular expression
    date_pattern = re.compile(r'(\d{1,2})/(\d{1,2})/(\d{2,4})')

    #search for the pattern in the text objects text
    date = re.search(date_pattern, text.text)

    if date:
        #.group() ensures it is returned as a string and not an object
        return date.group()
    else:
        return None

#check which date to use and return that date
def select_date(date1, date2):

    #if any of the dates are null return the other one or manual
    if date1 is None and (date2 is None or not last_30(date2)):
        return 'MANUAL'
    elif date1 is None:
        result = convert_date(date2)
    elif date2 is None:
        result = convert_date(date1)
    #elif bottom date is in the last 30 days then bottom date
    elif last_30(date2):
        result = convert_date(date2)
    #else top date
    else:
        result = convert_date(date1)

    if result is None:
        return 'MANUAL'

    return result

#convert string to date object
def convert_date(date):

    #possible formats the date could come in
    formats = ['%m/%d/%Y', '%m/%d/%y']

    #iterate through formats
    for format in formats:
        try:
            #try the date with a format
            return datetime.strptime(date, format).date()

        except ValueError:
            #if it doesnt work, try it with the next format
            continue

    return None

#calculate if a date was within the last 30 days
def last_30(date):

    #convert to a date object
    date_object = convert_date(date)

    if date_object is None:
        return False

    #calculate today
    today = datetime.today().date()

    #determine 30 days ago
    last30 = today - timedelta(days=30)

    #if the passed date is within today and 30 days ago, return true
    return last30 <= date_object <= today

def apply_style(date, cell):

    red_fill = PatternFill(start_color = 'FF0000', end_color = 'FF0000', fill_type = 'solid')
    purple_text = Font(color = '800080')
    
    if date == 'MANUAL' or date == 'LOAD ERROR':
        cell.fill = red_fill    
    else:
        cell.font = purple_text

if __name__ == "__main__":
    main()

#profit
    
